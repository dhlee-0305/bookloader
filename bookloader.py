import configparser
import logging
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pymysql
from openpyxl import load_workbook

from kakao import search_book

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("bookloader.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ── 도서 목록 설정 ────────────────────────────────────────────────
BOOK_COLUMN_MAP = {
    "책 이름": "title",
    "작가": "author",
    "출판사": "publisher",
    "구입 날짜": "purchaseDate",
    "상태": "status",
}

BOOK_STATUS_MAP = {
    "기부": "DONATED",
    "판매": "SOLD",
}

INSERT_BOOK_SQL = """
    INSERT IGNORE INTO books (title, author, publisher, purchaseDate, updatedAt, status, isbn, coverUrl)
    VALUES (%(title)s, %(author)s, %(publisher)s, %(purchaseDate)s, %(purchaseDate)s, %(status)s, %(isbn)s, %(coverUrl)s)
"""

# ── 독서 기록 설정 ────────────────────────────────────────────────
# 시트별 필수 컬럼 (보류목록은 '년' 미사용)
READING_LOG_COLUMNS = {
    "독서목록": ["년", "대현", "문선"],
    "보류목록": ["대현", "문선"],
}

USER_MAP = {
    "대현": "dhlee.0305@gmail.com",
    "문선": "yurina99@gmail.com",
}

# 시트명 → readStatus 매핑
READING_STATUS_MAP = {
    "독서목록": "READ",
    "보류목록": "EXCLUDED",
}

INSERT_READING_LOG_SQL = """
    INSERT INTO reading_logs (bookId, createdAt, updatedAt, userName, readStatus)
    VALUES (%(bookId)s, %(createdAt)s, %(updatedAt)s, %(userName)s, %(readStatus)s)
"""


# ── 공통 유틸 ──────────────────────────────────────────────────────
def load_config(config_path: str = "config.ini") -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    if not config.read(config_path, encoding="utf-8"):
        raise FileNotFoundError(f"설정 파일을 찾을 수 없습니다: {config_path}")
    return config


def get_db_connection(config: configparser.ConfigParser) -> pymysql.connections.Connection:
    db_cfg = config["database"]
    return pymysql.connect(
        host=db_cfg["host"],
        port=int(db_cfg.get("port", 3306)),
        user=db_cfg["user"],
        password=db_cfg["password"],
        database=db_cfg["database"],
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )


# ── 도서 목록 처리 ────────────────────────────────────────────────
def read_book_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """도서 목록 시트 읽기 (헤더: 2행)"""
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=1, dtype=str)

    missing = [col for col in BOOK_COLUMN_MAP if col not in df_raw.columns]
    if missing:
        raise ValueError(f"[{sheet_name}] 시트에 필수 컬럼이 없습니다: {missing}")

    df = df_raw[list(BOOK_COLUMN_MAP.keys())].rename(columns=BOOK_COLUMN_MAP)

    # isbn, coverUrl 컬럼 추가 (없거나 NaN이면 빈 문자열로 초기화)
    for col in ("isbn", "coverUrl"):
        df[col] = df_raw[col].fillna("").astype(str) if col in df_raw.columns else ""

    df = df.dropna(subset=list(BOOK_COLUMN_MAP.values()), how="all")
    df["purchaseDate"] = (
        df["purchaseDate"]
        .str.replace(r"\s+", "", regex=True)
        .str.replace(r"00$", "01", regex=True)
        .fillna("1999-01-01")
    )
    df["status"] = df["status"].map(BOOK_STATUS_MAP).fillna("OWNED")
    df = df.where(pd.notna(df), None)
    return df


def enrich_books_with_kakao(df: pd.DataFrame, config: configparser.ConfigParser) -> pd.DataFrame:
    """isbn 또는 coverUrl이 없는 도서를 카카오 도서 검색 API로 보완합니다.

    - 첫 번째 검색 결과의 isbn / thumbnail을 사용합니다.
    - 검색 결과가 없으면 빈 문자열('')을 저장합니다.
    """
    needs_enrich_mask = (
        df["isbn"].isna() | (df["isbn"] == "") |
        df["coverUrl"].isna() | (df["coverUrl"] == "")
    )
    target_indices = df.index[needs_enrich_mask].tolist()

    if not target_indices:
        logger.info("카카오 API 보완 대상 없음")
        return df

    logger.info("카카오 API로 도서 정보 보완 대상: %d건", len(target_indices))

    for idx in target_indices:
        row = df.loc[idx]
        isbn_val = row.get("isbn")
        cover_val = row.get("coverUrl")

        # 둘 중 값이 하나라도 없으면 API 호출 대상. 둘 다 존재하면 건너뜀
        if (len(isbn_val) > 5) and (len(cover_val) > 5):
            logger.info("이미 isbn과 coverUrl이 존재하여 건너뜀 - title: '%s', isbn: '%s', coverUrl: '%s'", row.get("title"), isbn_val, cover_val)
            continue

        title     = str(row.get("title") or "").strip()
        author    = str(row.get("author") or "").strip()
        publisher = str(row.get("publisher") or "").strip()

        if not any([title, author, publisher]):
            logger.warning("검색어를 구성할 수 없어 건너뜀 (idx=%s)", idx)
            continue

        logger.info(
            "카카오 API 호출 - title: '%s' | author: '%s' | publisher: '%s'",
            title, author, publisher,
        )
        try:
            books = search_book(config, title=title, author=author, publisher=publisher)
            if books:
                first = books[0]
                new_isbn  = first.get("isbn", "") or ""
                new_cover = first.get("thumbnail", "") or ""
                if not isbn_val:
                    df.loc[idx, "isbn"] = new_isbn
                if not cover_val:
                    df.loc[idx, "coverUrl"] = new_cover
                logger.info(
                    "카카오 검색 성공 - title: '%s' | isbn: '%s' | coverUrl: '%s'",
                    title, new_isbn, new_cover,
                )
            else:
                if not isbn_val:
                    df.loc[idx, "isbn"] = ""
                if not cover_val:
                    df.loc[idx, "coverUrl"] = ""
                logger.warning(
                    "카카오 검색 결과 없음 - title: '%s' | author: '%s' | publisher: '%s'",
                    title, author, publisher,
                )
        except Exception as e:
            logger.error(
                "카카오 검색 오류 - title: '%s' | author: '%s' | publisher: '%s' | error: %s",
                title, author, publisher, e,
            )
            if not isbn_val:
                df.loc[idx, "isbn"] = ""
            if not cover_val:
                df.loc[idx, "coverUrl"] = ""

    return df


def update_excel_book_details(file_path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """isbn과 coverUrl을 원본 엑셀 파일에 업데이트합니다.

    header=1 로 읽은 DataFrame의 index i 는 엑셀 1-based 행 번호 i+3 에 해당합니다.
    (row 1 = 빈 행, row 2 = 헤더, row 3 = 데이터 첫 행)
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 헤더 행(2행, 1-based)에서 isbn, coverUrl 컬럼 번호 탐색
    header_row_num = 2
    col_indices: dict[str, int] = {}
    for cell in ws[header_row_num]:
        if cell.value in ("isbn", "coverUrl"):
            col_indices[cell.value] = cell.column

    if not col_indices:
        logger.warning("[%s] 엑셀에서 isbn/coverUrl 컬럼을 찾을 수 없어 업데이트를 건너뜁니다.", sheet_name)
        wb.close()
        return

    updated = 0
    for i, row in df.iterrows():
        excel_row = i + 3  # pandas index → 엑셀 1-based 행
        for col_name, col_idx in col_indices.items():
            value = row.get(col_name)
            ws.cell(row=excel_row, column=col_idx, value=value if value is not None else "")
            updated += 1

    wb.save(file_path)
    logger.info("[%s] 엑셀 isbn/coverUrl 업데이트 완료 (%d 셀)", sheet_name, updated)


def insert_books(cursor, df: pd.DataFrame) -> tuple[int, int]:
    inserted, skipped = 0, 0
    for _, row in df.iterrows():
        # pandas NaN(float)은 MySQL에 전달 불가 → None으로 변환
        record = {k: None if isinstance(v, float) and pd.isna(v) else v for k, v in row.to_dict().items()}
        cursor.execute(INSERT_BOOK_SQL, record)
        if cursor.rowcount == 1:
            inserted += 1
        else:
            skipped += 1
            logger.debug("중복으로 건너뜀 - title: %s", record.get("title"))
    return inserted, skipped


def process_book_sheets(cursor, excel_path: str, sheet_names: list[str], config: configparser.ConfigParser):
    for sheet in sheet_names:
        logger.info("=== 도서 목록 시트 처리: [%s] ===", sheet)
        try:
            df = read_book_sheet(excel_path, sheet)
            logger.info("[%s] 읽은 행 수: %d", sheet, len(df))

            # isbn/coverUrl 보완 (카카오 API)
            df = enrich_books_with_kakao(df, config)

            # 원본 엑셀 파일에 isbn/coverUrl 업데이트
            update_excel_book_details(excel_path, sheet, df)

            inserted, skipped = insert_books(cursor, df)
            logger.info("[%s] 저장 완료 - 삽입: %d건, 중복 무시: %d건", sheet, inserted, skipped)
        except ValueError as e:
            logger.error(e)
        except Exception as e:
            logger.error("[%s] 처리 중 오류 발생: %s", sheet, e)


# ── 독서 기록 처리 ────────────────────────────────────────────────
def load_books_cache(cursor) -> dict[str, int]:
    """books 테이블의 title → id 매핑을 캐시로 반환"""
    cursor.execute("SELECT id, title FROM books")
    cache = {row["title"]: row["id"] for row in cursor.fetchall() if row["title"]}
    logger.info("books 캐시 로드 완료: %d건", len(cache))
    return cache


def read_reading_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """독서/보류 목록 시트 읽기 (헤더: 2행)"""
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=1, dtype=str)

    required_cols = READING_LOG_COLUMNS[sheet_name]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"[{sheet_name}] 시트에 필수 컬럼이 없습니다: {missing}")

    df = df[required_cols].dropna(how="all")
    df = df.where(pd.notna(df), None)
    return df


def insert_reading_logs(cursor, df: pd.DataFrame, sheet_name: str, books_cache: dict[str, int]) -> tuple[int, int]:
    read_status = READING_STATUS_MAP[sheet_name]
    inserted, errors = 0, 0

    for _, row in df.iterrows():
        # createdAt 결정
        if sheet_name == "보류목록":
            # 보류목록: 현재 시간 사용
            created_at = datetime.now()
        else:
            # 독서목록: '년' 컬럼 → date(YYYY-01-01), 변환 불가 시 1999-01-01
            year_val = row.get("년")
            if year_val is None:
                continue
            try:
                created_at = date(int(float(year_val)), 1, 1)
            except (ValueError, TypeError):
                logger.warning(
                    "[%s] 년도 변환 불가 (값: '%s') - createdAt을 1999-01-01로 설정",
                    sheet_name, year_val,
                )
                created_at = date(1999, 1, 1)

        # 사용자별(대현, 문선) 처리
        for col_name, user_name in USER_MAP.items():
            title = row.get(col_name)
            if not title or str(title).strip().lower() == "nan":
                continue

            book_id = books_cache.get(title)
            if book_id is None:
                logger.error(
                    "[%s] books 테이블에서 도서를 찾을 수 없습니다 - title: '%s' (사용자: %s)",
                    sheet_name, title, user_name,
                )
                errors += 1
                continue

            cursor.execute(INSERT_READING_LOG_SQL, {
                "bookId": book_id,
                "createdAt": created_at,
                "updatedAt": datetime.now(),
                "userName": user_name,
                "readStatus": read_status,
            })
            inserted += 1

    return inserted, errors


def process_reading_log_sheets(cursor, excel_path: str, sheet_names: list[str], books_cache: dict[str, int]):
    for sheet in sheet_names:
        logger.info("=== 독서 기록 시트 처리: [%s] ===", sheet)
        try:
            df = read_reading_sheet(excel_path, sheet)
            logger.info("[%s] 읽은 행 수: %d", sheet, len(df))
            inserted, errors = insert_reading_logs(cursor, df, sheet, books_cache)
            logger.info("[%s] 저장 완료 - 삽입: %d건, 오류: %d건", sheet, inserted, errors)
        except ValueError as e:
            logger.error(e)
        except Exception as e:
            logger.error("[%s] 처리 중 오류 발생: %s", sheet, e)


# ── 메인 ──────────────────────────────────────────────────────────
def main():
    config = load_config()

    excel_path = config["excel"]["file_path"]
    if not Path(excel_path).exists():
        logger.error("엑셀 파일을 찾을 수 없습니다: %s", excel_path)
        sys.exit(1)

    book_sheets = [s.strip() for s in config["excel"]["sheets"].split(",")]
    reading_log_sheets = [s.strip() for s in config["reading_logs"]["sheets"].split(",")]

    conn = get_db_connection(config)
    try:
        with conn.cursor() as cursor:
            # 1단계: 도서 목록 저장 (isbn/coverUrl 보완 포함)
            process_book_sheets(cursor, excel_path, book_sheets, config)
            conn.commit()
            logger.info("도서 목록 저장 완료")

            # 2단계: books 캐시 로드 후 독서 기록 저장
            books_cache = load_books_cache(cursor)
            process_reading_log_sheets(cursor, excel_path, reading_log_sheets, books_cache)
            conn.commit()
            logger.info("독서 기록 저장 완료")

        logger.info("=== 전체 처리 완료 ===")

    except Exception as e:
        conn.rollback()
        logger.error("DB 오류로 롤백 처리: %s", e)
        sys.exit(1)

    finally:
        conn.close()


if __name__ == "__main__":
    main()
